# src/helpers/data_profiling/profile_excel_files.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart, LineChart, ScatterChart, PieChart, 
    AreaChart, BubbleChart, RadarChart, DoughnutChart
)
import json

def detect_charts_in_sheet(worksheet):
    """
    Detects if there are any chart objects in the given Excel worksheet.
    
    Args:
    worksheet (Worksheet): An openpyxl worksheet object.

    Returns:
    bool: True if there are charts present in the worksheet, False otherwise.
    """
    # Check for any chart objects in the sheet
    chart_types = (BarChart, LineChart, ScatterChart, PieChart, 
                   AreaChart, BubbleChart, RadarChart, DoughnutChart)
    return any(isinstance(drawing, chart_types) for drawing in worksheet._charts)


def detect_formulas_in_sheet(worksheet):
    """
    Detects if there are any formulas in the given Excel worksheet and provides examples.

    Args:
    worksheet (Worksheet): An openpyxl worksheet object.

    Returns:
    dict: Information about whether the sheet contains formulas and columns with formula examples.
    """
    contains_formulas = False
    columns_with_formulas = {}

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Check if the cell has a formula
                contains_formulas = True
                col_letter = cell.column_letter
                # Record one example of the formula for this column
                if col_letter not in columns_with_formulas:
                    columns_with_formulas[col_letter] = cell.value  # Example formula

    return {
        'contains_formulas': contains_formulas,
        'columns_with_formulas': columns_with_formulas
    }

def profile_excel_file(file_path):
    """
    Profiles an Excel file to document its schema, presence of charts, formulas, and missing values.
    
    Args:
    file_path (str): The path to the Excel file to be profiled.

    Returns:
    dict: A dictionary containing the schema, chart presence, formulas, and missing value information for each sheet.
    """
    # Load the workbook to access charts
    workbook = load_workbook(file_path, data_only=False)  # Load with formulas
    file_profile = {'filepath': file_path}

    for sheet_name in workbook.sheetnames:
        sheet_profile = {}
        worksheet = workbook[sheet_name]

        # Detect charts in the current sheet
        sheet_profile['charts'] = detect_charts_in_sheet(worksheet)

        # Detect formulas in the current sheet
        formulas_info = detect_formulas_in_sheet(worksheet)
        sheet_profile.update(formulas_info)

        # Use pandas to read the Excel file and load the specific sheet
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            # Extract schema information
            sheet_profile['columns'] = list(df.columns)
            sheet_profile['num_rows'] = len(df)
            sheet_profile['num_columns'] = len(df.columns)
            # Calculate % of missing values
            sheet_profile['missing_values_percentage'] = df.isna().mean().round(4).to_dict()  # Convert Series to dict
            # Extract sample data (10 rows: 1, 10, 20, etc.)
            sample_indices = list(range(0, min(len(df), 100), 10))
            sheet_profile['sample_data'] = df.iloc[sample_indices].fillna('').astype(str).to_dict(orient='records')  # Convert sample to JSON-serializable format

        except Exception as e:
            sheet_profile['error'] = str(e)
            print(f"Error reading sheet {sheet_name}: {e}")

        file_profile[sheet_name] = sheet_profile

    return file_profile

def profile_excel_files(directory_path, output_folder, exclude_folders=None):
    """
    Profiles all Excel files in the given directory and its subdirectories,
    excluding specified subfolders, and saves each profile result as a JSON file 
    in the output folder.
    
    Args:
    directory_path (str): The path to the directory containing Excel files.
    output_folder (str): The path to the output folder for saving JSON files.
    exclude_folders (list): A list of subfolder names to exclude from profiling.

    Returns:
    dict: A dictionary containing profiles of all Excel files processed.
    """
    if exclude_folders is None:
        exclude_folders = []

    os.makedirs(output_folder, exist_ok=True)

    all_profiles = {}

    for root, dirs, files in os.walk(directory_path):
        dirs[:] = [d for d in dirs if d not in exclude_folders]
        
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                file_path = os.path.join(root, file)
                print(f"Profiling {file_path}...")

                profile = profile_excel_file(file_path)
                all_profiles[file_path] = profile

                relative_path = os.path.relpath(root, directory_path)
                output_dir_path = os.path.join(output_folder, relative_path)
                os.makedirs(output_dir_path, exist_ok=True)
                output_file_name = os.path.splitext(file)[0] + '.json'
                output_file_path = os.path.join(output_dir_path, output_file_name)

                with open(output_file_path, 'w') as f:
                    json.dump(profile, f, indent=4)

    print(f"Profiled all Excel files successfully! JSON files are saved in '{output_folder}'.")
    return all_profiles
