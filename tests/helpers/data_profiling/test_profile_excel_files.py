import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import (BarChart)
from src.helpers.data_profiling.profile_excel_files import (
    detect_charts_in_sheet,
    detect_formulas_in_sheet,
    profile_excel_file,
    profile_excel_files
)

def test_detect_charts_in_sheet():
    # Create a mock worksheet with a chart
    wb = Workbook()
    ws = wb.active
    chart = BarChart()
    ws.add_chart(chart, "A1")

    # Test detect_charts_in_sheet function
    assert detect_charts_in_sheet(ws) == True

    # Remove chart and test again
    ws._charts.remove(chart)
    assert detect_charts_in_sheet(ws) == False

def test_detect_formulas_in_sheet():
    # Create a mock worksheet with formulas
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "=SUM(1, 2)"  # Example formula

    # Test detect_formulas_in_sheet function
    result = detect_formulas_in_sheet(ws)
    assert result['contains_formulas'] == True
    assert 'A' in result['columns_with_formulas']

    # Remove formulas and test again
    ws['A1'] = 5  # Replace formula with a number
    result = detect_formulas_in_sheet(ws)
    assert result['contains_formulas'] == False
    assert len(result['columns_with_formulas']) == 0

def test_profile_excel_file(tmp_path):
    # Create a mock Excel file
    file_path = tmp_path / "test.xlsx"
    df = pd.DataFrame({
        'A': [1, 2, 3],
        'B': [4, 5, 6],
        'C': [None, None, 9]
    })
    df.to_excel(file_path, index=False)

    # Test profile_excel_file function
    profile = profile_excel_file(file_path)
    assert 'test.xlsx' in str(profile['filepath'])  # Convert PosixPath to string
    assert 'Sheet1' in profile
    assert profile['Sheet1']['num_rows'] == 3
    assert profile['Sheet1']['num_columns'] == 3
    assert 'A' in profile['Sheet1']['columns']
    assert profile['Sheet1']['missing_values_percentage']['C'] == 0.6667

def test_profile_excel_files(tmp_path):
    # Create a temporary directory and Excel files
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    output_dir = tmp_path / "output"
    output_dir.mkdir()

    # Mock Excel files
    df = pd.DataFrame({
        'A': [1, 2, 3],
        'B': [4, 5, 6],
        'C': [None, None, 9]
    })
    excel_path = input_dir / "file1.xlsx"
    df.to_excel(excel_path, index=False)

    # Test profile_excel_files function
    profiles = profile_excel_files(str(input_dir), str(output_dir))
    assert len(profiles) == 1
    assert str(excel_path) in profiles  # Check for the full file path
    assert 'Sheet1' in profiles[str(excel_path)]
    assert profiles[str(excel_path)]['Sheet1']['num_rows'] == 3
    assert profiles[str(excel_path)]['Sheet1']['num_columns'] == 3
    assert os.path.exists(output_dir / "file1.json")

